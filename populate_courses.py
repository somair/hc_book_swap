import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hc_books.settings')
import django
django.setup()

from books.models import Subject, Course
from openpyxl import load_workbook

BOOKS_FILE_NAME = 'hc_books_1718'

def add_subjects(subjects, isbns):
    valid_books = []
    for index, cell in enumerate(subjects, start=1):
        if index == len(subjects):
            break
        if isbns[index].value is not None:
            valid_books.append(index)
            subject = Subject.objects.get_or_create(name=cell.value)[0]
            subject.save()
    return valid_books

def populate_courses():
    wb = load_workbook(filename=BOOKS_FILE_NAME+'.xlsx')
    ws = wb.active
    # Get columns from spreadsheet
    subjects = ws['A']
    isbns = ws['D']
    courses = ws['C']
    # List with indices of courses that have books
    valid_books = add_subjects(subjects, isbns)
    for index in valid_books:
        course = Course.objects.get_or_create(name=courses[index].value, subject=Subject.objects.get(name=subjects[index].value), teacher='N/A')[0]
        course.save()

if __name__ == '__main__':
    print("Populating courses...")
    populate_courses()